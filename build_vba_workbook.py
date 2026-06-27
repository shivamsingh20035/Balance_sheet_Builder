import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_openpyxl_template():
    print("Building styled Excel sheets and formulas using openpyxl...")
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    double_bottom = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # 1. Create worksheets
    ws_readme = wb.create_sheet(title="Home")
    ws_bd = wb.create_sheet(title="BD")
    ws_tb = wb.create_sheet(title="TB")
    ws_ae = wb.create_sheet(title="AE")
    ws_bs = wb.create_sheet(title="BS")
    ws_pl = wb.create_sheet(title="PL")
    ws_cf = wb.create_sheet(title="CF")
    ws_ageing = wb.create_sheet(title="Ageing")
    ws_n1 = wb.create_sheet(title="N1")
    ws_n2 = wb.create_sheet(title="N2")
    ws_n3 = wb.create_sheet(title="N3")
    ws_n4 = wb.create_sheet(title="N4")
    ws_n5 = wb.create_sheet(title="N5")
    ws_n6 = wb.create_sheet(title="N6")
    ws_tx = wb.create_sheet(title="TX")
    ws_val = wb.create_sheet(title="Validation")
    ws_coa = wb.create_sheet(title="Chart of Accounts")
    ws_drill = wb.create_sheet(title="Drilldown_View")
    ws_drill.sheet_state = "hidden"

    # Set gridlines visible
    for ws in wb.worksheets:
        ws.views.sheetView[0].showGridLines = True

    # Helper: apply button style
    def apply_button_style(cell):
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Helper: apply card style
    def apply_card_style(cell, is_val=False):
        if is_val:
            cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            cell.font = Font(name="Calibri", size=9, bold=True, color="555555")
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Helper: Link back to dashboard
    def add_home_link(ws_target):
        ws_target.cell(row=1, column=1, value='=HYPERLINK("#Home!A1", "<- Go to Dashboard")').font = Font(
            name="Calibri", bold=True, color="0000FF", underline="single"
        )

    # ----------------------------------------------------
    # 1. Home Dashboard Sheet
    # ----------------------------------------------------
    ws_readme.merge_cells("A2:H3")
    ws_readme["A2"] = "Balance Sheet Builder (AS) - Lite version"
    ws_readme["A2"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws_readme["A2"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_readme["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_readme.row_dimensions[2].height = 25
    ws_readme.row_dimensions[3].height = 25

    # Control Panel background fill
    for r in range(5, 24):
        for c in range(1, 9):
            ws_readme.cell(row=r, column=c).fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Left buttons
    left_buttons = [
        ("BD - Basic Details", "BD"),
        ("TB - Ledger Groupings", "TB"),
        ("AE - Adjustment Entries", "AE"),
        ("BS - Balance Sheet", "BS"),
        ("PL - Profit & Loss", "PL"),
        ("CF - Cash Flow", "CF"),
        ("TX - Tax Export", "TX"),
    ]
    for idx, (label, target) in enumerate(left_buttons):
        row_idx = 5 + (idx * 2)
        cell = ws_readme.cell(row=row_idx, column=2)
        cell.value = f'=HYPERLINK("#\'{target}\'!A1", "{label}")'
        apply_button_style(cell)
        ws_readme.row_dimensions[row_idx].height = 24

    # Middle buttons
    middle_buttons = [
        ("N1 - Accounting Policies", "N1"),
        ("N2 - Liabilities", "N2"),
        ("N3 - Fixed Assets", "N3"),
        ("N4 - Assets", "N4"),
        ("N5 - Income & Expense", "N5"),
        ("N6 - Other Notes", "N6"),
    ]
    for idx, (label, target) in enumerate(middle_buttons):
        row_idx = 5 + (idx * 2)
        cell = ws_readme.cell(row=row_idx, column=4)
        cell.value = f'=HYPERLINK("#\'{target}\'!A1", "{label}")'
        apply_button_style(cell)
        ws_readme.row_dimensions[row_idx].height = 24

    # KPI Cards
    ws_readme["F5"] = "Balance Sheet"
    ws_readme["F6"] = '=IF(Validation!D7="PASS", "RECONCILED", "UNRECONCILED")'
    apply_card_style(ws_readme["F5"], is_val=False)
    apply_card_style(ws_readme["F6"], is_val=True)

    ws_readme["G5"] = "Cash Flow"
    ws_readme["G6"] = '=IF(Validation!D8="PASS", "RECONCILED", "UNRECONCILED")'
    apply_card_style(ws_readme["G5"], is_val=False)
    apply_card_style(ws_readme["G6"], is_val=True)

    ws_readme["H5"] = "Notes Count"
    ws_readme["H6"] = "11 Active"
    apply_card_style(ws_readme["H5"], is_val=False)
    apply_card_style(ws_readme["H6"], is_val=True)

    ws_readme["F9"] = "Financial Year"
    ws_readme["F10"] = "='BD'!B5"
    apply_card_style(ws_readme["F9"], is_val=False)
    apply_card_style(ws_readme["F10"], is_val=True)

    ws_readme["G9"] = "Revenue"
    ws_readme["G10"] = '=IFERROR(VLOOKUP("3", \'PL\'!$A:$D, 3, FALSE), 0)'
    apply_card_style(ws_readme["G9"], is_val=False)
    apply_card_style(ws_readme["G10"], is_val=True)
    ws_readme["G10"].number_format = '₹#,##0.00'

    ws_readme["H9"] = "Profit (Loss)"
    ws_readme["H10"] = '=IFERROR(VLOOKUP("3", \'PL\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("4", \'PL\'!$A:$D, 3, FALSE), 0)'
    apply_card_style(ws_readme["H9"], is_val=False)
    apply_card_style(ws_readme["H10"], is_val=True)
    ws_readme["H10"].number_format = '₹#,##0.00'

    ws_readme.merge_cells("F13:F14")
    ws_readme["F13"] = "VBA Control Panel"
    ws_readme["F13"].font = Font(name="Calibri", size=10, bold=True, color="333333")
    ws_readme["F13"].alignment = Alignment(horizontal="center", vertical="center")

    ws_readme.merge_cells("G13:H13")
    ws_readme.merge_cells("G14:H14")
    ws_readme["G13"] = "Rounding Scale"
    ws_readme["G14"] = "='BD'!B7"
    apply_card_style(ws_readme["G13"], is_val=False)
    apply_card_style(ws_readme["G14"], is_val=True)

    ws_readme.column_dimensions["A"].width = 5
    ws_readme.column_dimensions["B"].width = 25
    ws_readme.column_dimensions["C"].width = 5
    ws_readme.column_dimensions["D"].width = 25
    ws_readme.column_dimensions["E"].width = 5
    ws_readme.column_dimensions["F"].width = 20
    ws_readme.column_dimensions["G"].width = 20
    ws_readme.column_dimensions["H"].width = 20

    # Tally Integration & Import Help Guide
    ws_readme.merge_cells("B25:H25")
    ws_readme["B25"] = "Tally Integration & Import Help Guide"
    ws_readme["B25"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_readme["B25"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_readme["B25"].alignment = Alignment(horizontal="center", vertical="center")
    ws_readme.row_dimensions[25].height = 24

    guide_steps = [
        ("Step 1: Export TB from Tally", "In Tally Prime or Tally ERP 9, navigate to 'Display More Reports' -> 'Trial Balance'."),
        ("Step 2: Excel Format", "Press Alt+E (Export) -> select format 'Excel (Spreadsheet)'. Ensure Ledger Name, Debit, and Credit columns are visible."),
        ("Step 3: Run Macro Import", "In this workbook's Home dashboard, click 'Import Trial Balance' and select the exported Tally file."),
        ("Step 4: Map New Accounts", "Go to the 'TB' sheet. Select Type, Primary Group, and Secondary Group dropdowns. The Node Code resolves automatically."),
        ("Step 5: View Statements", "Once mapped, all statements (Balance Sheet, P&L, Cash Flow, and Notes) will automatically populate and reconcile.")
    ]
    curr_guide_row = 26
    for step_title, step_desc in guide_steps:
        ws_readme.merge_cells(start_row=curr_guide_row, start_column=2, end_row=curr_guide_row, end_column=3)
        ws_readme.cell(row=curr_guide_row, column=2, value=step_title).font = Font(name="Calibri", size=9, bold=True, color="111111")
        ws_readme.cell(row=curr_guide_row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws_readme.cell(row=curr_guide_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        
        ws_readme.merge_cells(start_row=curr_guide_row, start_column=4, end_row=curr_guide_row, end_column=8)
        ws_readme.cell(row=curr_guide_row, column=4, value=step_desc).font = Font(name="Calibri", size=9, color="555555")
        ws_readme.cell(row=curr_guide_row, column=4).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws_readme.cell(row=curr_guide_row, column=4).alignment = Alignment(horizontal="left", vertical="center")
        
        for col_idx in range(2, 9):
            ws_readme.cell(row=curr_guide_row, column=col_idx).border = thin_border
            
        ws_readme.row_dimensions[curr_guide_row].height = 20
        curr_guide_row += 1

    # ----------------------------------------------------
    # 2. BD - Basic Details
    # ----------------------------------------------------
    add_home_link(ws_bd)
    ws_bd["A3"] = "Basic Details Configuration"
    ws_bd["A3"].font = Font(name="Calibri", size=14, bold=True)
    settings_data = [
        ("Parameter", "Setting Value"),
        ("Company Name", "Mock Enterprise Ltd."),
        ("Financial Year", "FY 2025-26"),
        ("Reporting Currency", "INR"),
        ("Rounding Type", "Unrounded"),
        ("Scale Factor", '=IF(B7="Thousands", 1000, IF(B7="Lakhs", 100000, IF(B7="Crores", 10000000, 1)))'),
        ("Entity Type", "Corporate (Schedule III)")
    ]
    for idx, (param, val) in enumerate(settings_data, start=3):
        ws_bd.cell(row=idx, column=1, value=param).font = Font(bold=True)
        ws_bd.cell(row=idx, column=2, value=val)
        ws_bd.cell(row=idx, column=1).border = thin_border
        ws_bd.cell(row=idx, column=2).border = thin_border
    
    ws_bd.column_dimensions["A"].width = 20
    ws_bd.column_dimensions["B"].width = 30

    # ----------------------------------------------------
    # 3. Chart of Accounts (COA)
    # ----------------------------------------------------
    nodes = [
        # BS - Equity and Liabilities
        {"code": "1", "name": "EQUITY AND LIABILITIES", "node_type": "EQUITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.1", "name": "Shareholders' Funds", "node_type": "EQUITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.1.1", "name": "Share Capital", "node_type": "EQUITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.1.2", "name": "Reserves and Surplus", "node_type": "EQUITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.2", "name": "Non-Current Liabilities", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.2.1", "name": "Long-term borrowings", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.3", "name": "Current Liabilities", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.3.1", "name": "Short-term borrowings", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.3.2", "name": "Trade payables", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.3.3", "name": "Other current liabilities", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        {"code": "1.3.4", "name": "Short-term provisions", "node_type": "LIABILITY", "statement_type": "BALANCE_SHEET"},
        # BS - Assets
        {"code": "2", "name": "ASSETS", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.1", "name": "Non-Current Assets", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.1.1", "name": "Property, Plant and Equipment and Intangible Assets", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.1.1.1", "name": "Property, Plant and Equipment", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.1.1.2", "name": "Intangible assets", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.1.2", "name": "Non-current investments", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2", "name": "Current Assets", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.1", "name": "Current investments", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.2", "name": "Inventories", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.3", "name": "Trade receivables", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.4", "name": "Cash and cash equivalents", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.5", "name": "Short-term loans and advances", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        {"code": "2.2.6", "name": "Other current assets", "node_type": "ASSET", "statement_type": "BALANCE_SHEET"},
        # PL - Income & Expense
        {"code": "3", "name": "INCOME", "node_type": "REVENUE", "statement_type": "PROFIT_LOSS"},
        {"code": "3.1", "name": "Revenue from operations", "node_type": "REVENUE", "statement_type": "PROFIT_LOSS"},
        {"code": "3.2", "name": "Other income", "node_type": "REVENUE", "statement_type": "PROFIT_LOSS"},
        {"code": "4", "name": "EXPENSES", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.1", "name": "Cost of materials consumed", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.2", "name": "Purchase of Stock-in-Trade", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.3", "name": "Employee benefits expense", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.4", "name": "Finance costs", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.5", "name": "Depreciation and amortization expense", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
        {"code": "4.6", "name": "Other expenses", "node_type": "EXPENSE", "statement_type": "PROFIT_LOSS"},
    ]

    ws_coa.append(["Node Code", "Node Name", "Type", "Statement"])
    ws_coa.cell(row=1, column=1).font = Font(bold=True)
    ws_coa.cell(row=1, column=2).font = Font(bold=True)
    ws_coa.cell(row=1, column=3).font = Font(bold=True)
    ws_coa.cell(row=1, column=4).font = Font(bold=True)
    for n in nodes:
        ws_coa.append([n["code"], n["name"], n["node_type"], n["statement_type"]])
    for row in ws_coa.iter_rows(min_row=1, max_row=len(nodes) + 1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # ----------------------------------------------------
    # 4. AE - Adjustment Entries
    # ----------------------------------------------------
    add_home_link(ws_ae)
    ws_ae["A3"] = "Post Adjustment Entries"
    ws_ae["A3"].font = Font(name="Calibri", size=14, bold=True)
    ws_ae.cell(row=5, column=1, value="GL Code")
    ws_ae.cell(row=5, column=2, value="GL Name")
    ws_ae.cell(row=5, column=3, value="Debit Amount")
    ws_ae.cell(row=5, column=4, value="Credit Amount")
    ws_ae.cell(row=5, column=5, value="Narration")
    for c in range(1, 6):
        ws_ae.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_ae.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for r in range(6, 56):
        ws_ae.cell(row=r, column=2, value=f'=IF(ISBLANK(A{r}), "", IFERROR(VLOOKUP(A{r}, \'TB\'!$A:$B, 2, FALSE), "Invalid GL"))')
        ws_ae.cell(row=r, column=3).number_format = '#,##0.00'
        ws_ae.cell(row=r, column=4).number_format = '#,##0.00'
        for c in range(1, 6):
            ws_ae.cell(row=r, column=c).border = thin_border

    # ----------------------------------------------------
    # 5. TB - Unified Ledger Groupings (3-Year Comparative)
    # ----------------------------------------------------
    add_home_link(ws_tb)
    
    # Title & Cards area
    ws_tb.merge_cells("A2:O2")
    ws_tb["A2"] = "Ledger Groupings"
    ws_tb["A2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_tb["A2"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_tb["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_tb.row_dimensions[2].height = 22

    # Headers structure
    ws_tb.merge_cells("C4:E4")
    ws_tb["C4"] = "31 March 2023"
    ws_tb.merge_cells("F4:H4")
    ws_tb["F4"] = "31 March 2024"
    ws_tb.merge_cells("I4:K4")
    ws_tb["I4"] = "31 March 2025"
    ws_tb.merge_cells("L4:O4")
    ws_tb["L4"] = "Groupings"
    
    for c_idx in [3, 6, 9, 12]:
        cell = ws_tb.cell(row=4, column=c_idx)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
    ws_tb.row_dimensions[4].height = 20

    sub_headers = [
        "GL Code", "GL Name", 
        "Books", "Adjustments", "Final", 
        "Books", "Adjustments", "Final", 
        "Books", "Adjustments", "Final", 
        "Type", "Primary Group", "Secondary Group", "Link"
    ]
    for idx, sh in enumerate(sub_headers, start=1):
        cell = ws_tb.cell(row=5, column=idx, value=sh)
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if idx > 2 else "left", vertical="center")
        cell.border = thin_border
    ws_tb.row_dimensions[5].height = 22

    # Mapped sample ledger items
    sample_gls = [
        # Code, Name, 2023 Books, 2024 Books, 2025 Books, Type, Primary, Secondary, Link
        ("1001", "Equity Share Capital", 0, -250000, -500000, "Equity", "Shareholders' Funds", "Share Capital", "1.1.1"),
        ("1002", "Retained Earnings", 0, 0, -107500, "Equity", "Shareholders' Funds", "Reserves and Surplus", "1.1.2"),
        ("2001", "Bank Term Loan", 0, -100000, -200000, "Liability", "Non-Current Liabilities", "Long-term borrowings", "1.2.1"),
        ("3001", "Trade Creditors", 0, -40000, -80000, "Liability", "Current Liabilities", "Trade payables", "1.3.2"),
        ("4001", "Plant and Machinery", 0, 200000, 400000, "Asset", "Non-Current Assets", "Property, Plant and Equipment", "2.1.1.1"),
        ("5001", "HDFC Bank A/c", 0, 227500, 562500, "Asset", "Current Assets", "Cash and cash equivalents", "2.2.4"),
        ("5002", "Accounts Receivable", 0, 30000, 60000, "Asset", "Current Assets", "Trade receivables", "2.2.3"),
        ("5003", "Inventory", 0, 40000, 80000, "Asset", "Current Assets", "Inventories", "2.2.2"),
        ("6001", "Sales Revenue", 0, -200000, -400000, "Income", "INCOME", "Revenue from operations", "3.1"),
        ("7001", "Employee Salaries", 0, 75000, 150000, "Expense", "EXPENSES", "Employee benefits expense", "4.3"),
        ("7002", "Interest Expense", 0, 7500, 15000, "Expense", "EXPENSES", "Finance costs", "4.4"),
        ("7004", "Office Rent", 0, 10000, 20000, "Expense", "EXPENSES", "Other expenses", "4.6"),
    ]

    for idx, row_data in enumerate(sample_gls, start=6):
        code, name, val23, val24, val25, g_type, g_pri, g_sec, g_lnk = row_data
        ws_tb.cell(row=idx, column=1, value=code)
        ws_tb.cell(row=idx, column=2, value=name)
        
        # 2023
        ws_tb.cell(row=idx, column=3, value=val23)
        ws_tb.cell(row=idx, column=4, value=0) # Adjustments
        ws_tb.cell(row=idx, column=5, value=f"=C{idx}+D{idx}")
        
        # 2024
        ws_tb.cell(row=idx, column=6, value=val24)
        ws_tb.cell(row=idx, column=7, value=0) # Adjustments
        ws_tb.cell(row=idx, column=8, value=f"=F{idx}+G{idx}")
        
        # 2025
        ws_tb.cell(row=idx, column=9, value=val25)
        ws_tb.cell(row=idx, column=10, value=f"=SUMIFS('AE'!$C$6:$C$55, 'AE'!$A$6:$A$55, A{idx}) - SUMIFS('AE'!$D$6:$D$55, 'AE'!$A$6:$A$55, A{idx})")
        ws_tb.cell(row=idx, column=11, value=f"=I{idx}+J{idx}")
        
        # Groupings
        ws_tb.cell(row=idx, column=12, value=g_type)
        ws_tb.cell(row=idx, column=13, value=g_pri)
        ws_tb.cell(row=idx, column=14, value=g_sec)
        ws_tb.cell(row=idx, column=15, value=f"=IFERROR(INDEX('Chart of Accounts'!$A$4:$A$100, MATCH(N{idx}, 'Chart of Accounts'!$B$4:$B$100, 0)), \"\")")

        for c in range(3, 12):
            ws_tb.cell(row=idx, column=c).number_format = '#,##0.00'
        for c in range(1, 16):
            ws_tb.cell(row=idx, column=c).border = thin_border
            ws_tb.cell(row=idx, column=c).font = Font(name="Calibri", size=10)
        ws_tb.row_dimensions[idx].height = 20

    tot_row = len(sample_gls) + 6
    ws_tb.cell(row=tot_row, column=1, value="Total").font = Font(bold=True)
    for c in range(3, 12):
        col_letter = get_column_letter(c)
        ws_tb.cell(row=tot_row, column=c, value=f"=SUM({col_letter}6:{col_letter}{tot_row-1})").font = Font(bold=True)
        ws_tb.cell(row=tot_row, column=c).number_format = '#,##0.00'
    for c in range(1, 16):
        ws_tb.cell(row=tot_row, column=c).border = double_bottom
    ws_tb.row_dimensions[tot_row].height = 22

    ws_tb.column_dimensions["A"].width = 10
    ws_tb.column_dimensions["B"].width = 25
    for c in range(3, 12):
        ws_tb.column_dimensions[get_column_letter(c)].width = 14
    ws_tb.column_dimensions["L"].width = 12
    ws_tb.column_dimensions["M"].width = 22
    ws_tb.column_dimensions["N"].width = 24
    ws_tb.column_dimensions["O"].width = 10

    # ----------------------------------------------------
    # 7. BS - Balance Sheet (Comparative CY and PY)
    # ----------------------------------------------------
    add_home_link(ws_bs)
    ws_bs.merge_cells("A2:D2")
    ws_bs["A2"] = "Mock Enterprise Ltd."
    ws_bs["A2"].font = Font(bold=True, size=12)
    ws_bs["A2"].alignment = Alignment(horizontal="center")
    ws_bs.merge_cells("A3:D3")
    ws_bs["A3"] = "Balance Sheet as at March 31, 2025"
    ws_bs["A3"].font = Font(italic=True, size=10)
    ws_bs["A3"].alignment = Alignment(horizontal="center")

    ws_bs.cell(row=5, column=1, value="Node Code")
    ws_bs.cell(row=5, column=2, value="Particulars")
    ws_bs.cell(row=5, column=3, value="31 March 2025")
    ws_bs.cell(row=5, column=4, value="31 March 2024")
    for c in range(1, 5):
        ws_bs.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_bs.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws_bs.cell(row=5, column=c).alignment = Alignment(horizontal="center" if c > 2 else "left", vertical="center")

    bs_nodes = [n for n in nodes if n["statement_type"] == "BALANCE_SHEET"]
    curr_row = 6
    bs_formulas = {
        "1": "=C7+C10+C12",
        "1.1": "=SUM(C8:C9)",
        "1.2": "=C11",
        "1.3": "=SUM(C13:C16)",
        "2": "=C18+C23",
        "2.1": "=C19+C22",
        "2.1.1": "=SUM(C20:C21)",
        "2.2": "=SUM(C24:C29)"
    }
    bs_particulars_formulas = {
        "1": '=IF(\'BD\'!$B$9="Non-Corporate", "CAPITAL AND LIABILITIES", "EQUITY AND LIABILITIES")',
        "1.1": '=IF(\'BD\'!$B$9="Non-Corporate", "Partners\' / Proprietor\'s Funds", "Shareholders\' Funds")',
        "1.1.1": '=IF(\'BD\'!$B$9="Non-Corporate", "Partners\' / Proprietor\'s Capital Account", "Share Capital")',
        "1.1.2": '=IF(\'BD\'!$B$9="Non-Corporate", "Reserves & Profit / Loss Balance", "Reserves and Surplus")'
    }
    for node in bs_nodes:
        code = node["code"]
        name = node["name"]
        is_parent = code in bs_formulas
        ws_bs.cell(row=curr_row, column=1, value=code)
        
        if code in bs_particulars_formulas:
            ws_bs.cell(row=curr_row, column=2, value=bs_particulars_formulas[code])
        else:
            ws_bs.cell(row=curr_row, column=2, value=name)
            
        amt_cell_cy = ws_bs.cell(row=curr_row, column=3)
        amt_cell_py = ws_bs.cell(row=curr_row, column=4)
        sign = -1 if node["node_type"] in ["LIABILITY", "EQUITY"] else 1
        sign_str = "-" if sign == -1 else ""

        if is_parent:
            ws_bs.cell(row=curr_row, column=2).font = Font(bold=True)
            
            # CY & PY Parent Formulas
            amt_cell_cy.value = bs_formulas[code]
            amt_cell_py.value = bs_formulas[code].replace("C", "D")
            
            amt_cell_cy.font = Font(bold=True)
            amt_cell_py.font = Font(bold=True)
            
            ws_bs.cell(row=curr_row, column=2).border = Border(top=Side(style='thin', color='000000'))
            amt_cell_cy.border = Border(top=Side(style='thin', color='000000'))
            amt_cell_py.border = Border(top=Side(style='thin', color='000000'))
            
            if len(code.split(".")) == 1:
                ws_bs.cell(row=curr_row, column=2).border = double_bottom
                amt_cell_cy.border = double_bottom
                amt_cell_py.border = double_bottom
        else:
            if code == "1.1.2":
                # Reserves & Surplus: link P&L net profit CY and PY
                amt_cell_cy.value = f"=({sign_str}SUMIFS('TB'!$K$6:$K$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8 + 'PL'!$C$6 - 'PL'!$C$9"
                amt_cell_py.value = f"=({sign_str}SUMIFS('TB'!$H$6:$H$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8 + 'PL'!$D$6 - 'PL'!$D$9"
            else:
                amt_cell_cy.value = f"=({sign_str}SUMIFS('TB'!$K$6:$K$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8"
                amt_cell_py.value = f"=({sign_str}SUMIFS('TB'!$H$6:$H$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8"
            
            ws_bs.cell(row=curr_row, column=2).alignment = Alignment(indent=1)

        amt_cell_cy.number_format = '#,##0.00'
        amt_cell_py.number_format = '#,##0.00'
        
        for c in range(1, 5):
            ws_bs.cell(row=curr_row, column=c).font = Font(name="Calibri", size=10)
        ws_bs.row_dimensions[curr_row].height = 20
        curr_row += 1

    ws_bs.column_dimensions["A"].width = 12
    ws_bs.column_dimensions["B"].width = 40
    ws_bs.column_dimensions["C"].width = 18
    ws_bs.column_dimensions["D"].width = 18
    ws_bs.column_dimensions["E"].width = 20

    # ----------------------------------------------------
    # 8. PL - Profit & Loss (Comparative CY and PY)
    # ----------------------------------------------------
    add_home_link(ws_pl)
    ws_pl.merge_cells("A2:D2")
    ws_pl["A2"] = "Mock Enterprise Ltd."
    ws_pl["A2"].font = Font(bold=True, size=12)
    ws_pl["A2"].alignment = Alignment(horizontal="center")
    ws_pl.merge_cells("A3:D3")
    ws_pl["A3"] = "Profit & Loss for the year ended March 31, 2025"
    ws_pl["A3"].font = Font(italic=True, size=10)
    ws_pl["A3"].alignment = Alignment(horizontal="center")

    ws_pl.cell(row=5, column=1, value="Node Code")
    ws_pl.cell(row=5, column=2, value="Particulars")
    ws_pl.cell(row=5, column=3, value="31 March 2025")
    ws_pl.cell(row=5, column=4, value="31 March 2024")
    for c in range(1, 5):
        ws_pl.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_pl.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws_pl.cell(row=5, column=c).alignment = Alignment(horizontal="center" if c > 2 else "left", vertical="center")

    pl_nodes = [n for n in nodes if n["statement_type"] == "PROFIT_LOSS"]
    curr_row = 6
    pl_formulas = {
        "3": "=SUM(C7:C8)",
        "4": "=SUM(C10:C15)"
    }
    for node in pl_nodes:
        code = node["code"]
        name = node["name"]
        is_parent = code in pl_formulas
        ws_pl.cell(row=curr_row, column=1, value=code)
        ws_pl.cell(row=curr_row, column=2, value=name)
        
        amt_cell_cy = ws_pl.cell(row=curr_row, column=3)
        amt_cell_py = ws_pl.cell(row=curr_row, column=4)
        sign = -1 if node["node_type"] == "REVENUE" else 1
        sign_str = "-" if sign == -1 else ""

        if is_parent:
            ws_pl.cell(row=curr_row, column=2).font = Font(bold=True)
            
            # CY & PY Parent Formulas
            amt_cell_cy.value = pl_formulas[code]
            amt_cell_py.value = pl_formulas[code].replace("C", "D")
            
            amt_cell_cy.font = Font(bold=True)
            amt_cell_py.font = Font(bold=True)
            
            ws_pl.cell(row=curr_row, column=2).border = Border(top=Side(style='thin', color='000000'))
            amt_cell_cy.border = Border(top=Side(style='thin', color='000000'))
            amt_cell_py.border = Border(top=Side(style='thin', color='000000'))
            
            if code in ["3", "4"]:
                ws_pl.cell(row=curr_row, column=2).border = double_bottom
                amt_cell_cy.border = double_bottom
                amt_cell_py.border = double_bottom
        else:
            amt_cell_cy.value = f"=({sign_str}SUMIFS('TB'!$K$6:$K$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8"
            amt_cell_py.value = f"=({sign_str}SUMIFS('TB'!$H$6:$H$500, 'TB'!$O$6:$O$500, A{curr_row})) / 'BD'!$B$8"
            
            ws_pl.cell(row=curr_row, column=2).alignment = Alignment(indent=1)

        amt_cell_cy.number_format = '#,##0.00'
        amt_cell_py.number_format = '#,##0.00'
        
        for c in range(1, 5):
            ws_pl.cell(row=curr_row, column=c).font = Font(name="Calibri", size=10)
        ws_pl.row_dimensions[curr_row].height = 20
        curr_row += 1

    ws_pl.column_dimensions["A"].width = 12
    ws_pl.column_dimensions["B"].width = 40
    ws_pl.column_dimensions["C"].width = 18
    ws_pl.column_dimensions["D"].width = 18
    ws_pl.column_dimensions["E"].width = 20

    # ----------------------------------------------------
    # 9. CF - Cash Flow
    # ----------------------------------------------------
    add_home_link(ws_cf)
    ws_cf.merge_cells("A2:C2")
    ws_cf["A2"] = "Indirect Cash Flow Statement"
    ws_cf["A2"].font = Font(bold=True, size=12)
    ws_cf["A2"].alignment = Alignment(horizontal="center")
    
    ws_cf.cell(row=4, column=1, value="Particulars")
    ws_cf.cell(row=4, column=2, value="Amount")
    ws_cf.cell(row=4, column=1).font = Font(bold=True, color="FFFFFF")
    ws_cf.cell(row=4, column=1).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_cf.cell(row=4, column=2).font = Font(bold=True, color="FFFFFF")
    ws_cf.cell(row=4, column=2).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    cf_lines = [
        ("Cash Flow from Operating Activities", ""),
        ("Net Profit Before Tax", '=IFERROR(VLOOKUP("3", \'PL\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("4", \'PL\'!$A:$D, 3, FALSE), 0)'),
        ("Adjustment: Depreciation and Amortization", '=IFERROR(VLOOKUP("4.5", \'PL\'!$A:$D, 3, FALSE), 0)'),
        ("Operating Profit before Working Capital Changes", "=SUM(B6:B7)"),
        ("Changes in Trade Payables", '=IFERROR(VLOOKUP("1.3.2", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("1.3.2", \'BS\'!$A:$D, 4, FALSE), 0)'),
        ("Changes in Trade Receivables", '=-(IFERROR(VLOOKUP("2.2.3", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("2.2.3", \'BS\'!$A:$D, 4, FALSE), 0))'),
        ("Changes in Inventories", '=-(IFERROR(VLOOKUP("2.2.2", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("2.2.2", \'BS\'!$A:$D, 4, FALSE), 0))'),
        ("Net Cash from Operating Activities", "=B8+B9+B10+B11"),
        ("Cash Flow from Investing Activities", ""),
        ("Purchase of Property, Plant and Equipment", '=-(IFERROR(VLOOKUP("2.1.1.1", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("2.1.1.1", \'BS\'!$A:$D, 4, FALSE), 0))'),
        ("Net Cash from Investing Activities", "=B14"),
        ("Cash Flow from Financing Activities", ""),
        ("Proceeds from Share Capital", '=IFERROR(VLOOKUP("1.1.1", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("1.1.1", \'BS\'!$A:$D, 4, FALSE), 0)'),
        ("Proceeds from Long-term Borrowings", '=IFERROR(VLOOKUP("1.2.1", \'BS\'!$A:$D, 3, FALSE), 0) - IFERROR(VLOOKUP("1.2.1", \'BS\'!$A:$D, 4, FALSE), 0)'),
        ("Net Cash from Financing Activities", "=B17+B18"),
        ("Net Increase / (Decrease) in Cash", "=B12+B15+B19"),
        ("Opening Cash & Cash Equivalents", '=IFERROR(VLOOKUP("2.2.4", \'BS\'!$A:$D, 4, FALSE), 0)'),
        ("Closing Cash & Cash Equivalents", "=B20+B21")
    ]

    for i, (particulars, formula) in enumerate(cf_lines, start=5):
        ws_cf.cell(row=i, column=1, value=particulars)
        amt_cell = ws_cf.cell(row=i, column=2, value=formula)
        amt_cell.number_format = '#,##0.00'
        if particulars.startswith("Cash Flow") or particulars.startswith("Net Cash") or particulars == "Net Increase / (Decrease) in Cash" or particulars.startswith("Closing"):
            ws_cf.cell(row=i, column=1).font = Font(bold=True)
            amt_cell.font = Font(bold=True)
        ws_cf.cell(row=i, column=1).border = thin_border
        ws_cf.cell(row=i, column=2).border = thin_border
        ws_cf.cell(row=i, column=1).font = Font(name="Calibri", size=10)
        amt_cell.font = Font(name="Calibri", size=10, bold=amt_cell.font.bold)

    ws_cf.column_dimensions["A"].width = 45
    ws_cf.column_dimensions["B"].width = 18

    # ----------------------------------------------------
    # 9b. Ageing - Ageing Schedules
    # ----------------------------------------------------
    add_home_link(ws_ageing)
    ws_ageing.merge_cells("A2:G2")
    ws_ageing["A2"] = "Ageing Schedules (Trade Receivables & Payables)"
    ws_ageing["A2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_ageing["A2"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_ageing["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ageing.row_dimensions[2].height = 25

    # --- Receivables Table ---
    ws_ageing.merge_cells("A4:G4")
    ws_ageing["A4"] = "Trade Receivables Ageing Schedule"
    ws_ageing["A4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_ageing["A4"].fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    ws_ageing["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws_ageing.row_dimensions[4].height = 20

    headers_rec = ["Particulars", "Less than 6 months", "6 months - 1 year", "1 - 2 years", "2 - 3 years", "More than 3 years", "Total"]
    for c_idx, h in enumerate(headers_rec, start=1):
        cell = ws_ageing.cell(row=5, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center", wrap_text=True)
    ws_ageing.row_dimensions[5].height = 24

    rows_rec = [
        "(i) Undisputed Trade Receivables - considered good",
        "(ii) Undisputed Trade Receivables - considered doubtful",
        "(iii) Disputed Trade Receivables - considered good",
        "(iv) Disputed Trade Receivables - considered doubtful"
    ]
    for r_offset, r_name in enumerate(rows_rec):
        r_idx = 6 + r_offset
        ws_ageing.cell(row=r_idx, column=1, value=r_name).font = Font(name="Calibri", size=10)
        for c_idx in range(2, 7):
            ws_ageing.cell(row=r_idx, column=c_idx, value=0.00).number_format = '#,##0.00'
        ws_ageing.cell(row=r_idx, column=7, value=f"=SUM(B{r_idx}:F{r_idx})").number_format = '#,##0.00'
        ws_ageing.cell(row=r_idx, column=7).font = Font(bold=True)
        for c_idx in range(1, 8):
            ws_ageing.cell(row=r_idx, column=c_idx).border = thin_border
        ws_ageing.row_dimensions[r_idx].height = 20

    # Total row
    ws_ageing.cell(row=10, column=1, value="Total Trade Receivables").font = Font(bold=True)
    for c_idx in range(2, 7):
        col_letter = get_column_letter(c_idx)
        cell = ws_ageing.cell(row=10, column=c_idx, value=f"=SUM({col_letter}6:{col_letter}9)")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
    ws_ageing.cell(row=10, column=7, value="=SUM(G6:G9)").font = Font(bold=True)
    ws_ageing.cell(row=10, column=7).number_format = '#,##0.00'
    for c_idx in range(1, 8):
        ws_ageing.cell(row=10, column=c_idx).border = double_bottom
    ws_ageing.row_dimensions[10].height = 20

    # Reconciliation rows for Receivables
    cell_bs_rec_lbl = ws_ageing.cell(row=11, column=1, value="Balance Sheet Control Total")
    cell_bs_rec_lbl.font = Font(name="Calibri", size=10, italic=True)
    cell_bs_rec_val = ws_ageing.cell(row=11, column=7, value="=IFERROR(VLOOKUP(\"2.2.3\", 'BS'!$A:$D, 3, FALSE), 0)")
    cell_bs_rec_val.font = Font(name="Calibri", size=10, italic=True)
    cell_bs_rec_val.number_format = '#,##0.00'
    ws_ageing.cell(row=11, column=1).border = thin_border
    ws_ageing.cell(row=11, column=7).border = thin_border
    ws_ageing.row_dimensions[11].height = 20

    cell_diff_rec_lbl = ws_ageing.cell(row=12, column=1, value="Unreconciled Difference")
    cell_diff_rec_lbl.font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    cell_diff_rec_val = ws_ageing.cell(row=12, column=7, value="=G10-G11")
    cell_diff_rec_val.font = Font(name="Calibri", size=10, bold=True)
    cell_diff_rec_val.number_format = '#,##0.00;[Red]-#,##0.00;"0.00"'
    ws_ageing.cell(row=12, column=1).border = thin_border
    ws_ageing.cell(row=12, column=7).border = thin_border
    ws_ageing.row_dimensions[12].height = 20

    # --- Payables Table ---
    ws_ageing.merge_cells("A14:G14")
    ws_ageing["A14"] = "Trade Payables Ageing Schedule"
    ws_ageing["A14"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_ageing["A14"].fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    ws_ageing["A14"].alignment = Alignment(horizontal="left", vertical="center")
    ws_ageing.row_dimensions[14].height = 20

    headers_pay = ["Particulars", "Less than 1 year", "1 - 2 years", "2 - 3 years", "More than 3 years", "", "Total"]
    for c_idx, h in enumerate(headers_pay, start=1):
        if h != "":
            cell = ws_ageing.cell(row=15, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center", wrap_text=True)
        else:
            cell = ws_ageing.cell(row=15, column=c_idx)
            cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    ws_ageing.row_dimensions[15].height = 24

    rows_pay = [
        "(i) MSME",
        "(ii) Others",
        "(iii) Disputed dues - MSME",
        "(iv) Disputed dues - Others"
    ]
    for r_offset, r_name in enumerate(rows_pay):
        r_idx = 16 + r_offset
        ws_ageing.cell(row=r_idx, column=1, value=r_name).font = Font(name="Calibri", size=10)
        for c_idx in range(2, 6):
            ws_ageing.cell(row=r_idx, column=c_idx, value=0.00).number_format = '#,##0.00'
        ws_ageing.cell(row=r_idx, column=6, value="-").alignment = Alignment(horizontal="center")
        ws_ageing.cell(row=r_idx, column=7, value=f"=SUM(B{r_idx}:E{r_idx})").number_format = '#,##0.00'
        ws_ageing.cell(row=r_idx, column=7).font = Font(bold=True)
        for c_idx in range(1, 8):
            ws_ageing.cell(row=r_idx, column=c_idx).border = thin_border
        ws_ageing.row_dimensions[r_idx].height = 20

    # Total row
    ws_ageing.cell(row=20, column=1, value="Total Trade Payables").font = Font(bold=True)
    for c_idx in range(2, 6):
        col_letter = get_column_letter(c_idx)
        cell = ws_ageing.cell(row=20, column=c_idx, value=f"=SUM({col_letter}16:{col_letter}19)")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
    ws_ageing.cell(row=20, column=6, value="-").alignment = Alignment(horizontal="center")
    ws_ageing.cell(row=20, column=7, value="=SUM(G16:G19)").font = Font(bold=True)
    ws_ageing.cell(row=20, column=7).number_format = '#,##0.00'
    for c_idx in range(1, 8):
        ws_ageing.cell(row=20, column=c_idx).border = double_bottom
    ws_ageing.row_dimensions[20].height = 20

    # Reconciliation rows for Payables
    cell_bs_pay_lbl = ws_ageing.cell(row=21, column=1, value="Balance Sheet Control Total")
    cell_bs_pay_lbl.font = Font(name="Calibri", size=10, italic=True)
    cell_bs_pay_val = ws_ageing.cell(row=21, column=7, value="=IFERROR(VLOOKUP(\"1.3.2\", 'BS'!$A:$D, 3, FALSE), 0)")
    cell_bs_pay_val.font = Font(name="Calibri", size=10, italic=True)
    cell_bs_pay_val.number_format = '#,##0.00'
    ws_ageing.cell(row=21, column=1).border = thin_border
    ws_ageing.cell(row=21, column=7).border = thin_border
    ws_ageing.row_dimensions[21].height = 20

    cell_diff_pay_lbl = ws_ageing.cell(row=22, column=1, value="Unreconciled Difference")
    cell_diff_pay_lbl.font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    cell_diff_pay_val = ws_ageing.cell(row=22, column=7, value="=G20-G21")
    cell_diff_pay_val.font = Font(name="Calibri", size=10, bold=True)
    cell_diff_pay_val.number_format = '#,##0.00;[Red]-#,##0.00;"0.00"'
    ws_ageing.cell(row=22, column=1).border = thin_border
    ws_ageing.cell(row=22, column=7).border = thin_border
    ws_ageing.row_dimensions[22].height = 20

    ws_ageing.column_dimensions["A"].width = 45
    ws_ageing.column_dimensions["B"].width = 15
    ws_ageing.column_dimensions["C"].width = 15
    ws_ageing.column_dimensions["D"].width = 15
    ws_ageing.column_dimensions["E"].width = 15
    ws_ageing.column_dimensions["F"].width = 10
    ws_ageing.column_dimensions["G"].width = 15

    # ----------------------------------------------------
    # 10. Notes Sheets N1-N6
    # ----------------------------------------------------
    # N1 Policies
    add_home_link(ws_n1)
    ws_n1["A3"] = "Note 1: Significant Accounting Policies"
    ws_n1["A3"].font = Font(bold=True, size=12)
    policies = [
        "1. Basis of Preparation:",
        "The financial statements are prepared under historical cost convention on an accrual basis of accounting in accordance with the standard Indian Accounting Standards (AS) and the provisions of the Companies Act, 2013.",
        "",
        "2. Property, Plant and Equipment (PPE):",
        "PPE are stated at cost of acquisition less accumulated depreciation. Depreciation is calculated using the Straight Line Method (SLM) based on the useful lives of assets prescribed under Schedule II of the Companies Act, 2013.",
        "",
        "3. Revenue Recognition:",
        "Revenue is recognized when control of the goods or services is transferred to the customer at an amount that reflects the consideration to which the company expects to be entitled."
    ]
    for idx, line in enumerate(policies, start=5):
        ws_n1.cell(row=idx, column=1, value=line)
        ws_n1.cell(row=idx, column=1).font = Font(name="Calibri", size=10)

    # Reusable note writer function
    gl_accounts_list = [{"gl_code": g[0], "gl_name": g[1]} for g in sample_gls]
    gl_node_mappings = {g[0]: g[8] for g in sample_gls}
    
    # N2 - Liabilities Note
    add_home_link(ws_n2)
    r_liab = 4
    r_liab = ExcelEngine_write_note(ws_n2, "1.1.1", "Share Capital", 2, r_liab, gl_accounts_list, gl_node_mappings, thin_border)
    r_liab = ExcelEngine_write_note(ws_n2, "1.1.2", "Reserves and Surplus", 3, r_liab, gl_accounts_list, gl_node_mappings, thin_border)
    r_liab = ExcelEngine_write_note(ws_n2, "1.3.2", "Trade Payables", 4, r_liab, gl_accounts_list, gl_node_mappings, thin_border)

    # N3 - Fixed Assets
    add_home_link(ws_n3)
    ExcelEngine_write_note(ws_n3, "2.1.1.1", "Property, Plant and Equipment", 5, 4, gl_accounts_list, gl_node_mappings, thin_border)

    # N4 - Assets
    add_home_link(ws_n4)
    r_assets = 4
    r_assets = ExcelEngine_write_note(ws_n4, "2.2.3", "Trade Receivables", 6, r_assets, gl_accounts_list, gl_node_mappings, thin_border)
    r_assets = ExcelEngine_write_note(ws_n4, "2.2.4", "Cash and Cash Equivalents", 7, r_assets, gl_accounts_list, gl_node_mappings, thin_border)

    # N5 - Income & Expense
    add_home_link(ws_n5)
    r_inc = 4
    r_inc = ExcelEngine_write_note(ws_n5, "3.1", "Revenue from Operations", 8, r_inc, gl_accounts_list, gl_node_mappings, thin_border)
    r_inc = ExcelEngine_write_note(ws_n5, "4.3", "Employee Benefit Expenses", 9, r_inc, gl_accounts_list, gl_node_mappings, thin_border)

    # N6 - Ratios
    add_home_link(ws_n6)
    ws_n6["A3"] = "Note 11: Analytical Ratios"
    ws_n6["A3"].font = Font(bold=True, size=12)
    
    ws_n6.cell(row=5, column=1, value="Ratio Name")
    ws_n6.cell(row=5, column=2, value="Formula / Details")
    ws_n6.cell(row=5, column=3, value="Current Value")
    
    for c in range(1, 4):
        ws_n6.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_n6.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    ratios = [
        ("Current Ratio", "Current Assets / Current Liabilities", '=IFERROR(IFERROR(VLOOKUP("2.2", \'BS\'!$A:$D, 3, FALSE), 0)/IFERROR(VLOOKUP("1.3", \'BS\'!$A:$D, 3, FALSE), 1), 0)'),
        ("Debt to Equity Ratio", "Long-term Borrowings / Shareholders' Funds", '=IFERROR(IFERROR(VLOOKUP("1.2", \'BS\'!$A:$D, 3, FALSE), 0)/IFERROR(VLOOKUP("1.1", \'BS\'!$A:$D, 3, FALSE), 1), 0)'),
    ]
    for idx, (lbl, f_dtl, formula) in enumerate(ratios, start=6):
        ws_n6.cell(row=idx, column=1, value=lbl)
        ws_n6.cell(row=idx, column=2, value=f_dtl)
        cell = ws_n6.cell(row=idx, column=3, value=formula)
        cell.number_format = '0.00'
        for c in range(1, 4):
            ws_n6.cell(row=idx, column=c).border = thin_border
            ws_n6.cell(row=idx, column=c).font = Font(name="Calibri", size=10)

    ws_n6.column_dimensions["A"].width = 25
    ws_n6.column_dimensions["B"].width = 35
    ws_n6.column_dimensions["C"].width = 15

    # ----------------------------------------------------
    # 9c. TX - Income Tax Export
    # ----------------------------------------------------
    add_home_link(ws_tx)
    ws_tx.merge_cells("A2:C2")
    ws_tx["A2"] = "Income Tax Return - Balance Sheet and Profit & Loss"
    ws_tx["A2"].font = Font(name="Calibri", size=14, bold=True)
    ws_tx.merge_cells("A3:C3")
    ws_tx["A3"] = "Year ended March 31, 2025"
    ws_tx["A3"].font = Font(name="Calibri", size=10, italic=True)
    ws_tx.merge_cells("A4:C4")
    ws_tx["A4"] = "All amounts in INR Unrounded, unless otherwise stated"
    ws_tx["A4"].font = Font(name="Calibri", size=8, italic=True, color="555555")

    ws_tx.cell(row=5, column=1, value="Field Name")
    ws_tx.cell(row=5, column=2, value="Amount CY")
    ws_tx.cell(row=5, column=3, value="Amount PY")
    for c in range(1, 4):
        ws_tx.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_tx.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws_tx.cell(row=5, column=c).border = thin_border

    tx_items = [
        # Name, Node, Sheet
        ("Share Capital", "1.1.1", "BS"),
        ("Reserves and Surplus", "1.1.2", "BS"),
        ("Long-term Borrowings", "1.2.1", "BS"),
        ("Short-term Borrowings", "1.3.1", "BS"),
        ("Trade Payables", "1.3.2", "BS"),
        ("Other Current Liabilities", "1.3.3", "BS"),
        ("Short-term Provisions", "1.3.4", "BS"),
        ("Property, Plant and Equipment", "2.1.1.1", "BS"),
        ("Intangible Assets", "2.1.1.2", "BS"),
        ("Non-Current Investments", "2.1.2", "BS"),
        ("Trade Receivables", "2.2.3", "BS"),
        ("Cash and Cash Equivalents", "2.2.4", "BS"),
        ("Short-term Loans and Advances", "2.2.5", "BS"),
        ("Other Current Assets", "2.2.6", "BS"),
        ("Revenue from Operations", "3.1", "PL"),
        ("Other Income", "3.2", "PL"),
        ("Cost of Materials Consumed", "4.1", "PL"),
        ("Purchase of Stock-in-Trade", "4.2", "PL"),
        ("Employee Benefits Expense", "4.3", "PL"),
        ("Finance Costs", "4.4", "PL"),
        ("Depreciation and Amortization", "4.5", "PL"),
        ("Other Expenses", "4.6", "PL")
    ]

    for idx, (name, code, sh) in enumerate(tx_items, start=6):
        ws_tx.cell(row=idx, column=1, value=name)
        ws_tx.cell(row=idx, column=2, value=f'=IFERROR(VLOOKUP("{code}", {sh}!$A:$D, 3, FALSE), 0)')
        ws_tx.cell(row=idx, column=3, value=f'=IFERROR(VLOOKUP("{code}", {sh}!$A:$D, 4, FALSE), 0)')
        
        ws_tx.cell(row=idx, column=2).number_format = '#,##0.00'
        ws_tx.cell(row=idx, column=3).number_format = '#,##0.00'
        for c in range(1, 4):
            ws_tx.cell(row=idx, column=c).border = thin_border
            ws_tx.cell(row=idx, column=c).font = Font(name="Calibri", size=10)

    ws_tx.column_dimensions["A"].width = 30
    ws_tx.column_dimensions["B"].width = 20
    ws_tx.column_dimensions["C"].width = 20
    ws_tx.column_dimensions["E"].width = 20

    # ----------------------------------------------------
    # 11. Validation Sheet
    # ----------------------------------------------------
    add_home_link(ws_val)
    ws_val["A3"] = "Diagnostic Validation Controls"
    ws_val["A3"].font = Font(name="Calibri", size=14, bold=True)
    
    ws_val.cell(row=5, column=1, value="Diagnostic Description")
    ws_val.cell(row=5, column=2, value="Actual Value 1")
    ws_val.cell(row=5, column=3, value="Actual Value 2")
    ws_val.cell(row=5, column=4, value="Status")
    for c in range(1, 5):
        ws_val.cell(row=5, column=c).font = Font(bold=True, color="FFFFFF")
        ws_val.cell(row=5, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    ws_val.cell(row=6, column=1, value="Trial Balance Debits = Credits")
    ws_val.cell(row=6, column=2, value="=SUMIF('TB'!$K$6:$K$500, \">0\")")
    ws_val.cell(row=6, column=3, value="=-SUMIF('TB'!$K$6:$K$500, \"<0\")")
    ws_val.cell(row=6, column=4, value='=IF(ROUND(B6-C6,2)=0, "PASS", "FAIL")')

    ws_val.cell(row=7, column=1, value="Total Assets = Total Equity & Liabilities")
    ws_val.cell(row=7, column=2, value='=IFERROR(VLOOKUP("2", \'BS\'!$A:$D, 3, FALSE), 0)')
    ws_val.cell(row=7, column=3, value='=IFERROR(VLOOKUP("1", \'BS\'!$A:$D, 3, FALSE), 0)')
    ws_val.cell(row=7, column=4, value='=IF(ROUND(B7-C7,2)=0, "PASS", "FAIL")')

    ws_val.cell(row=8, column=1, value="Cash Flow Reconciled (Opening + Change = Closing)")
    ws_val.cell(row=8, column=2, value='=IFERROR(VLOOKUP("Closing Cash & Cash Equivalents", \'CF\'!$A:$B, 2, FALSE), 0)')
    ws_val.cell(row=8, column=3, value='=IFERROR(VLOOKUP("2.2.4", \'BS\'!$A:$D, 3, FALSE), 0)')
    ws_val.cell(row=8, column=4, value='=IF(ROUND(B8-C8,2)=0, "PASS", "FAIL")')

    for r in range(6, 9):
        ws_val.cell(row=r, column=2).number_format = '#,##0.00'
        ws_val.cell(row=r, column=3).number_format = '#,##0.00'
        for c in range(1, 5):
            ws_val.cell(row=r, column=c).border = thin_border
            ws_val.cell(row=r, column=c).font = Font(name="Calibri", size=10)
            if c == 4:
                ws_val.cell(row=r, column=c).font = Font(bold=True)

    ws_val.column_dimensions["A"].width = 45
    ws_val.column_dimensions["B"].width = 18
    ws_val.column_dimensions["C"].width = 18
    ws_val.column_dimensions["D"].width = 15

    # ----------------------------------------------------
    # Save openpyxl template
    # ----------------------------------------------------
    temp_filename = "temp_sheet_openpyxl.xlsx"
    if os.path.exists(temp_filename):
        os.remove(temp_filename)
    wb.save(temp_filename)
    print("Temporary spreadsheet created successfully!")
    return temp_filename

def ExcelEngine_write_note(ws, node_code, node_name, note_num, row_start, gl_accounts, mapping_dict, thin_border):
    ws.cell(row=row_start, column=1, value=f"Note {note_num}: {node_name}").font = Font(bold=True, size=11)
    row_start += 1
    
    ws.cell(row=row_start, column=1, value="GL Code")
    ws.cell(row=row_start, column=2, value="GL Particulars")
    ws.cell(row=row_start, column=3, value="Amount")
    
    for c in range(1, 4):
        ws.cell(row=row_start, column=c).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row_start, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    row_start += 1
    
    matching_gls = []
    for gl in gl_accounts:
        gl_code = gl["gl_code"]
        node_path = mapping_dict.get(gl_code, "")
        if node_path == node_code or node_path.startswith(node_code + "."):
            matching_gls.append(gl)
            
    start_data_row = row_start
    for gl in matching_gls:
        gl_code = gl["gl_code"]
        ws.cell(row=row_start, column=1, value=gl_code)
        ws.cell(row=row_start, column=2, value=gl["gl_name"])
        ws.cell(
            row=row_start, 
            column=3, 
            value=f'=IFERROR(VLOOKUP(A{row_start}, \'TB\'!$A:$O, 11, FALSE), 0) / \'BD\'!$B$8'
        )
        ws.cell(row=row_start, column=3).number_format = '#,##0.00'
        for c in range(1, 4):
            ws.cell(row=row_start, column=c).border = thin_border
            ws.cell(row=row_start, column=c).font = Font(name="Calibri", size=10)
        row_start += 1
        
    if not matching_gls:
        ws.cell(row=row_start, column=1, value="-")
        ws.cell(row=row_start, column=2, value="No GL accounts mapped")
        ws.cell(row=row_start, column=3, value=0.00)
        ws.cell(row=row_start, column=3).number_format = '#,##0.00'
        for c in range(1, 4):
            ws.cell(row=row_start, column=c).border = thin_border
            ws.cell(row=row_start, column=c).font = Font(name="Calibri", size=10)
        row_start += 1
        
    end_data_row = row_start - 1
    ws.cell(row=row_start, column=2, value=f"Total {node_name}").font = Font(bold=True)
    total_cell = ws.cell(row=row_start, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    
    double_bottom = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    for c in range(1, 4):
        ws.cell(row=row_start, column=c).border = double_bottom
        ws.cell(row=row_start, column=c).font = Font(name="Calibri", size=10, bold=True)
        
    return row_start + 2

def inject_vba_and_save_xlsm(temp_xlsx_path):
    print("Connecting to Microsoft Excel COM to inject VBA macros and save as .xlsm...")
    try:
        import win32com.client
    except ImportError:
        print("\n[WARNING]: pywin32 (win32com) is not installed.")
        print("Falling back. Saving template as 'Balance_Sheet_Builder_Template.xlsx'.")
        if os.path.exists("Balance_Sheet_Builder_Template.xlsx"):
            os.remove("Balance_Sheet_Builder_Template.xlsx")
        os.rename(temp_xlsx_path, "Balance_Sheet_Builder_Template.xlsx")
        return False

    import winreg
    reg_path = r"Software\Microsoft\Office\16.0\Excel\Security"
    original_val = 0
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path, 0, winreg.KEY_READ)
        original_val, _ = winreg.QueryValueEx(key, "AccessVBOM")
        winreg.CloseKey(key)
    except Exception:
        pass

    try:
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, reg_path)
        winreg.SetValueEx(key, "AccessVBOM", 0, winreg.REG_DWORD, 1)
        winreg.CloseKey(key)
        print("Temporarily enabled programmatic access to VBA project object model.")
    except Exception as e:
        print(f"Warning: Could not set registry AccessVBOM: {e}")

    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_xlsx = os.path.abspath(temp_xlsx_path)
        wb = excel.Workbooks.Open(abs_xlsx)
        
        base_dir = os.path.dirname(os.path.abspath(__file__))
        nav_bas = os.path.join(base_dir, "vba_code", "Navigation.bas")
        import_bas = os.path.join(base_dir, "vba_code", "ImportTB.bas")
        export_bas = os.path.join(base_dir, "vba_code", "ExportPDF.bas")
        export_tax_bas = os.path.join(base_dir, "vba_code", "ExportTax.bas")
        
        audit_bas = os.path.join(base_dir, "vba_code", "AuditTool.bas")
        
        # Import VBA Modules
        print("Importing Navigation.bas...")
        wb.VBProject.VBComponents.Import(nav_bas)
        print("Importing ImportTB.bas...")
        wb.VBProject.VBComponents.Import(import_bas)
        print("Importing ExportPDF.bas...")
        wb.VBProject.VBComponents.Import(export_bas)
        print("Importing ExportTax.bas...")
        wb.VBProject.VBComponents.Import(export_tax_bas)
        print("Importing AuditTool.bas...")
        wb.VBProject.VBComponents.Import(audit_bas)
        
        # Inject double-click event listener into ThisWorkbook
        print("Injecting double-click event listener into ThisWorkbook...")
        this_wb_code = (
            "Private Sub Workbook_SheetBeforeDoubleClick(ByVal Sh As Object, ByVal Target As Range, Cancel As Boolean)\n"
            "    On Error Resume Next\n"
            "    Call AuditTool.TraceCellToLedgers(Sh, Target, Cancel)\n"
            "End Sub\n"
        )
        wb.VBProject.VBComponents("ThisWorkbook").CodeModule.AddFromString(this_wb_code)
        
        # Dashboard sheet
        ws_readme = wb.Sheets("Home")
        
        # Navigation & Control buttons positioning
        col_f = ws_readme.Columns(6)
        col_f_width = col_f.Width
        col_f_left = col_f.Left

        # Button 1: Show Toolbar in cell F15:F16
        print("Adding Show Toolbar button...")
        cell_1 = ws_readme.Cells(15, 6)
        btn_toggle = ws_readme.Buttons().Add(col_f_left + 5, cell_1.Top, col_f_width - 10, cell_1.Height * 1.8)
        btn_toggle.OnAction = "ToggleToolbar"
        btn_toggle.Characters.Text = "Show/Hide Toolbar"

        # Button 2: Import Trial Balance in cell G15:G16 (Column G)
        print("Adding Import Trial Balance button...")
        col_g = ws_readme.Columns(7)
        col_g_width = col_g.Width
        col_g_left = col_g.Left
        cell_2 = ws_readme.Cells(15, 7)
        btn_import = ws_readme.Buttons().Add(col_g_left + 5, cell_2.Top, col_g_width - 10, cell_2.Height * 1.8)
        btn_import.OnAction = "ImportTrialBalance"
        btn_import.Characters.Text = "Import Trial Balance"

        # Button 3: Export to PDF in cell H15:H16 (Column H)
        print("Adding Export PDF button...")
        col_h = ws_readme.Columns(8)
        col_h_width = col_h.Width
        col_h_left = col_h.Left
        cell_3 = ws_readme.Cells(15, 8)
        btn_export = ws_readme.Buttons().Add(col_h_left + 5, cell_3.Top, col_h_width - 10, cell_3.Height * 1.8)
        btn_export.OnAction = "ExportStatementsToPDF"
        btn_export.Characters.Text = "Export to PDF"

        # Button 4: Roll Forward Year in cell G18:G19
        print("Adding Roll Forward Year button...")
        cell_4 = ws_readme.Cells(18, 7)
        btn_rollover = ws_readme.Buttons().Add(col_g_left + 5, cell_4.Top, col_g_width * 1.5, cell_4.Height * 1.8)
        btn_rollover.OnAction = "RollForwardYear"
        btn_rollover.Characters.Text = "Roll Forward Year"

        # ----------------------------------------------------
        # Add Page-Fitting Buttons on BS & PL
        # ----------------------------------------------------
        for sheet_name in ["BS", "PL"]:
            ws_sheet = wb.Sheets(sheet_name)
            col_e = ws_sheet.Columns(5) # Column E
            col_e_width = col_e.Width
            col_e_left = col_e.Left
            
            # Fit on Two Pages at Cell E2
            cell_fit2 = ws_sheet.Cells(2, 5)
            btn_fit2 = ws_sheet.Buttons().Add(col_e_left + 5, cell_fit2.Top, col_e_width - 10, cell_fit2.Height * 1.8)
            btn_fit2.OnAction = "FitToTwoPages"
            btn_fit2.Characters.Text = "Fit on Two Pages"
            
            # Fit on One Page at Cell E4
            cell_fit1 = ws_sheet.Cells(4, 5)
            btn_fit1 = ws_sheet.Buttons().Add(col_e_left + 5, cell_fit1.Top, col_e_width - 10, cell_fit1.Height * 1.8)
            btn_fit1.OnAction = "FitToOnePage"
            btn_fit1.Characters.Text = "Fit on One Page"

        # ----------------------------------------------------
        # Add CompuTax / Winman Export Buttons on TX
        # ----------------------------------------------------
        ws_tx = wb.Sheets("TX")
        col_tx_e = ws_tx.Columns(5) # Column E
        col_tx_e_width = col_tx_e.Width
        col_tx_e_left = col_tx_e.Left
        
        # CompuTax Export at cell E2
        cell_ct = ws_tx.Cells(2, 5)
        btn_ct = ws_tx.Buttons().Add(col_tx_e_left + 5, cell_ct.Top, col_tx_e_width - 10, cell_ct.Height * 1.8)
        btn_ct.OnAction = "ExportToCompuTax"
        btn_ct.Characters.Text = "CompuTax Export"
        
        # Winman Export at cell E4
        cell_wm = ws_tx.Cells(4, 5)
        btn_wm = ws_tx.Buttons().Add(col_tx_e_left + 5, cell_wm.Top, col_tx_e_width - 10, cell_wm.Height * 1.8)
        btn_wm.OnAction = "ExportToWinman"
        btn_wm.Characters.Text = "Winman Export"
        
        out_xlsm = os.path.join(base_dir, "Balance_Sheet_Builder_Macro.xlsm")
        wb.SaveAs(out_xlsm, FileFormat=52)
        wb.Close(SaveChanges=True)
        print(f"\n[SUCCESS]: Compiled macro-enabled workbook generated at:\n{out_xlsm}")
        
        if os.path.exists(temp_xlsx_path):
            os.remove(temp_xlsx_path)
        return True
        
    except Exception as e:
        print(f"\n[ERROR]: VBA compilation failed: {e}")
        if 'wb' in locals():
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if os.path.exists("Balance_Sheet_Builder_Template.xlsx"):
            os.remove("Balance_Sheet_Builder_Template.xlsx")
        try:
            os.rename(temp_xlsx_path, "Balance_Sheet_Builder_Template.xlsx")
        except Exception:
            pass
        return False
        
    finally:
        try:
            key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, reg_path)
            winreg.SetValueEx(key, "AccessVBOM", 0, winreg.REG_DWORD, original_val)
            winreg.CloseKey(key)
            print("Restored original programmatic access setting in registry.")
        except Exception:
            pass
        if excel:
            excel.Quit()

if __name__ == "__main__":
    temp_path = build_openpyxl_template()
    success = inject_vba_and_save_xlsm(temp_path)
